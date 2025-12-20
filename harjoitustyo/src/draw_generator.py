import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from draw import Draw
from database_connection import get_database_connection
from web_search import get_rating

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')

def get_pool_draw(class_name, draw: Draw):
    """
    Gets the draw for the given class
    Args:
        class_name: name of the class in string format
    Returns:
        dictionary of all the pools containing the players
        from the given class
    """
    return draw.draw_for_class(class_name)

def get_match_schedule(num_players):
    """
    Returns the playing order depending on the amount of players
    Args:
        num_players: the amount of players in a pool
    Returns:
        A list of tuples containing the order
    """
    if num_players == 3:
        return [(1, 3, 2), (1, 2, 3), (2, 3, 1)]
    if num_players == 4:
        return [
            (1, 3, 4), (2, 4, 3), (1, 2, 4),
            (3, 4, 2), (1, 4, 3), (2, 3, 1)
        ]
    if num_players == 5:
        return [
            (1, 5, 4), (2, 3, 3), (3, 5, 2), (1, 4, 3), (2, 5, 1),
            (1, 3, 2), (3, 4, 5), (1, 2, 4), (4, 5, 1), (2, 3, 5)
        ]
    return []

def format_cell(ws, row, col, value, font=None, alignment=None, border=None):
    """
    Determines the cell format and applies styles to a specific cell in a worksheet.
    Args:
        ws: worksheet object where the cell is located.
        row: the row number of the cell
        col: the column number of the cell
        value: the value to be written into the cell
        font: the font style to apply to the cell
        alignment: the alignment settings for the cell
        border: the border style to apply to the cell
    Returns:
        Cell: The cell object with the applied value and formatting.
    """
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell

def create_competition_excel(class_list, draw, output_filename="Competition_Draw.xlsx"):
    """
    Writes the draw into an excel file
    Args:
        class_list: list of all the classes in the competition
        output_filename: name of the outputted excel file
    Returns:
        None
    """
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for class_name in class_list:
        ws = wb.create_sheet(title=class_name)
        pool_data = get_pool_draw(class_name, draw)

        ws.cell(row=2, column=4, value="Competition").font = BOLD_FONT
        ws.cell(row=3, column=4, value=class_name).font = BOLD_FONT
        ws.cell(row=4, column=4, value="Pvm")
        ws.cell(row=4, column=6, value="Klo")

        current_row = 7

        for pool_name, players in pool_data.items():
            num_players = len(players)

            headers = [
                (4, "Rating"), (5, pool_name), (6, "Seura"),
                (7, "Voitot"), (8, "Erät"), (9, "Pisteet"), (10, "Sija")
            ]

            for col, text in headers:
                format_cell(ws, current_row, col, text, font=BOLD_FONT,
                             alignment=CENTER_ALIGN, border=THIN_BORDER)

            current_row += 1

            for idx, player in enumerate(players, 1):
                format_cell(ws, current_row, 3,
                            idx, alignment=CENTER_ALIGN)
                format_cell(ws, current_row, 4, player['rating'],
                            alignment=CENTER_ALIGN, border=THIN_BORDER)
                format_cell(ws, current_row, 5, player['name'],
                            alignment=LEFT_ALIGN, border=THIN_BORDER)
                format_cell(ws, current_row, 6, player['club'],
                            alignment=LEFT_ALIGN, border=THIN_BORDER)

                for c in range(7, 11):
                    format_cell(ws, current_row, c, "", border=THIN_BORDER)

                current_row += 1

            current_row += 2

            match_headers = ["1. erä", "2. erä", "3. erä", "4. erä", "5. erä", "Ottelu", "Tuomari"]
            for i, header in enumerate(match_headers):
                col_idx = 6 + i
                format_cell(ws, current_row, col_idx, header, font=BOLD_FONT,
                            alignment=CENTER_ALIGN, border=THIN_BORDER)

            current_row += 1

            schedule = get_match_schedule(num_players)

            for p1, p2, umpire in schedule:
                format_cell(ws, current_row, 5, f"{p1}-{p2}", font=BOLD_FONT,
                            alignment=CENTER_ALIGN, border=THIN_BORDER)

                for c in range(6, 11):
                    format_cell(ws, current_row, c, "", border=THIN_BORDER)

                format_cell(ws, current_row, 11, "", border=THIN_BORDER)

                format_cell(ws, current_row, 12, umpire, alignment=CENTER_ALIGN, border=THIN_BORDER)

                current_row += 1

            current_row += 3

        ws.column_dimensions['C'].width = 4
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        for col_char in ['G', 'H', 'I', 'J', 'K', 'L']:
            ws.column_dimensions[col_char].width = 8

    wb.save(output_filename)
    print(f"File saved as {output_filename}")

def generate(file_name, date):
    """
    Runs the whole process of generating the draw excel file
    Args:
        file_name: name of the entry excel file
        date: rating date used in the competition
    Returns:
        None
    """
    connection = get_database_connection()
    cursor = connection.cursor()
    date = date.split('.')
    date.reverse()
    date = '-'.join(date)
    get_rating(date, connection)
    draw = Draw(file_name, connection)
    if not draw.success:
        return False
    cursor.execute('SELECT class FROM Entries GROUP BY class')
    rows = cursor.fetchall()
    classes = [row[0] for row in rows]
    create_competition_excel(classes, draw)
    return True